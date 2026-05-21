import random
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework_simplejwt.tokens import RefreshToken
from django.utils import timezone
from django.db import models
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from django.contrib.auth.models import User
from django.core.mail import EmailMessage
from django.db import transaction
from django.conf import settings

from .utils import generate_document_pdf
from .models import Client, Vehicule, Reparation, Stock, Facture, LigneTravail, LignePiece, MouvementCaisse, Devis, MaintenancePredictive, Appointment, NotificationClient, Avis, NotificationStaff, MouvementStock
from .serializers import (
    ClientSerializer, VehiculeSerializer, ReparationSerializer, 
    StockSerializer, UserSerializer, MiniVehiculeSerializer,
    FactureSerializer, LigneTravailSerializer, LignePieceSerializer,
    MouvementCaisseSerializer, DevisSerializer, MaintenancePredictiveSerializer,
    AppointmentSerializer, NotificationClientSerializer, AvisSerializer,
    NotificationStaffSerializer
)

class IsDirecteur(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.user.is_superuser:
            return True
        return hasattr(request.user, 'profile') and request.user.profile.role == 'DIRECTEUR'

class IsStaffMember(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.user.is_superuser:
            return True
        return hasattr(request.user, 'profile') and request.user.profile.role in ['DIRECTEUR', 'SECRETAIRE']

from rest_framework.views import APIView

class GlobalSearchView(APIView):
    """
    GET /api/search/?q=<terme>
    Retourne jusqu'à 5 clients, 5 véhicules et 5 réparations correspondant au terme.
    """
    permission_classes = [IsStaffMember]

    def get(self, request):
        q = request.query_params.get('q', '').strip()
        if len(q) < 2:
            return Response({'clients': [], 'vehicules': [], 'reparations': []})

        clients = Client.objects.filter(
            models.Q(nom__icontains=q) |
            models.Q(prenoms__icontains=q) |
            models.Q(contact__icontains=q)
        )[:5]

        vehicules = Vehicule.objects.filter(
            models.Q(immatriculation__icontains=q) |
            models.Q(marque__icontains=q) |
            models.Q(modele__icontains=q)
        ).select_related('client')[:5]

        reparations = Reparation.objects.filter(
            models.Q(vehicule__immatriculation__icontains=q) |
            models.Q(description__icontains=q) |
            models.Q(categorie__icontains=q)
        ).select_related('vehicule__client')[:5]

        return Response({
            'clients': [
                {'id': c.id, 'nom': c.nom, 'prenoms': c.prenoms, 'contact': c.contact}
                for c in clients
            ],
            'vehicules': [
                {'id': v.id, 'immatriculation': v.immatriculation, 'marque': v.marque,
                 'modele': v.modele, 'client_nom': f"{v.client.nom} {v.client.prenoms}"}
                for v in vehicules
            ],
            'reparations': [
                {'id': r.id, 'description': r.description[:60], 'statut': r.statut,
                 'immatriculation': r.vehicule.immatriculation,
                 'client_nom': f"{r.vehicule.client.nom} {r.vehicule.client.prenoms}"}
                for r in reparations
            ],
        })

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all().order_by('-date_creation')
    serializer_class = ClientSerializer

class AppointmentViewSet(viewsets.ModelViewSet):
    queryset = Appointment.objects.all().order_by('-date_rdv')
    serializer_class = AppointmentSerializer
    permission_classes = [permissions.AllowAny]

    def perform_create(self, serializer):
        appointment = serializer.save()
        
        # Notification Staff
        NotificationStaff.objects.create(
            type='NOUVEAU_RDV',
            message=f"Nouveau rendez-vous pris par {appointment.nom_client_public or appointment.client} pour le {appointment.date_rdv.strftime('%d/%m/%Y %H:%M')}."
        )

        # Créer une notification pour le client (si lié) ou logguer la demande
        if appointment.client:
            NotificationClient.objects.create(
                client=appointment.client,
                type='RDV_CONFIRME',
                message=f"Votre demande de rendez-vous pour le {appointment.date_rdv.strftime('%d/%m/%Y à %H:%M')} a été enregistrée. Nous vous contacterons pour confirmation."
            )

class NotificationClientViewSet(viewsets.ModelViewSet):
    queryset = NotificationClient.objects.all().order_by('-date_envoi')
    serializer_class = NotificationClientSerializer

class AvisViewSet(viewsets.ModelViewSet):
    queryset = Avis.objects.all().order_by('-date_creation')
    serializer_class = AvisSerializer

    def get_permissions(self):
        if self.action in ['public_list', 'create', 'submit_avis']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['get'])
    def public_list(self, request):
        avis = Avis.objects.filter(affiche_public=True).order_by('-date_creation')[:6]
        serializer = self.get_serializer(avis, many=True)
        return Response(serializer.data)

class VehiculeViewSet(viewsets.ModelViewSet):
    queryset = Vehicule.objects.all().order_by('-date_creation')
    serializer_class = VehiculeSerializer

    @action(detail=True, methods=['get'])
    def historique(self, request, pk=None):
        vehicule = self.get_object()
        reparations = vehicule.reparations.all().order_by('-date_creation')
        serializer = ReparationSerializer(reparations, many=True)
        return Response(serializer.data)

class ReparationViewSet(viewsets.ModelViewSet):
    queryset = Reparation.objects.all().order_by('-date_creation')
    serializer_class = ReparationSerializer

    def partial_update(self, request, *args, **kwargs):
        """Force partial=True pour les requêtes PATCH (statut, progression...)."""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    def perform_update(self, serializer):
        reparation = serializer.save()
        # Si la réparation passe à TERMINE, on notifie le client
        if reparation.statut == 'TERMINE':
            client = reparation.vehicule.client
            immat = reparation.vehicule.immatriculation
            marque = reparation.vehicule.marque
            modele = reparation.vehicule.modele

            # 1. Notification en base de données (espace client)
            try:
                NotificationClient.objects.create(
                    client=client,
                    type='REPARATION_TERMINEE',
                    message=f"Votre véhicule {immat} est prêt ! Vous pouvez passer le récupérer au garage LUXEL-G à Parakou."
                )
            except Exception as e:
                print(f"[WARNING] Impossible de créer la notification en base : {e}")

            # 2. Email immédiat au client
            if client.email:
                try:
                    subject = f"✅ Votre véhicule {immat} est prêt — LUXEL-G"
                    body = (
                        f"Bonjour {client.prenoms} {client.nom},\n\n"
                        f"Nous avons le plaisir de vous informer que votre véhicule "
                        f"{marque} {modele} ({immat}) a été pris en charge et est désormais prêt à être récupéré.\n\n"
                        f"📍 Adresse : Garage LUXEL-G, Okedama, Parakou, Bénin\n"
                        f"📞 Contact : +229 01 92 62 98 60\n\n"
                        f"Merci de votre confiance.\n\n"
                        f"Cordialement,\n"
                        f"L'équipe LUXEL-G — Luxury Élégance Garage"
                    )
                    email = EmailMessage(subject, body, to=[client.email])
                    email.send()
                    print(f"[INFO] Email de notification envoyé à {client.email} pour {immat}.")
                except Exception as e:
                    print(f"[WARNING] Échec de l'envoi de l'email de notification : {e}")
            else:
                print(f"[INFO] Pas d'email pour le client {client.nom} — notification email ignorée.")


    def perform_create(self, serializer):
        reparation = serializer.save()
        if reparation.categorie == 'Maintenance Routine':
            km_prochain = reparation.kilometrage + 10000
            date_prochaine = timezone.now() + timezone.timedelta(days=180)
            MaintenancePredictive.objects.create(
                vehicule=reparation.vehicule,
                type_maintenance='VIDANGE',
                km_derniere_prestation=reparation.kilometrage,
                km_prochain_prevu=km_prochain,
                date_prochaine_prevue=date_prochaine
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = Workbook()
        ws = wb.active
        ws.append(['Date Entrée', 'OR #', 'Immatriculation', 'Marque', 'Modèle', 'Client', 'Description', 'Kms Entrée', 'Statut'])
        for r in self.get_queryset():
            ws.append([r.date_creation.strftime('%d/%m/%Y'), f"OR-{r.id:04d}", r.vehicule.immatriculation, r.vehicule.marque, r.vehicule.modele, r.vehicule.client.nom, r.description, r.kilometrage, r.get_statut_display()])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="registre_entrees_luxel_g.xlsx"'
        wb.save(response)
        return response

class FactureViewSet(viewsets.ModelViewSet):
    queryset = Facture.objects.all().order_by('-date_creation')
    serializer_class = FactureSerializer

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        facture = self.get_object()
        if facture.type == 'DEFINITIVE':
            return Response({'error': 'Facture déjà validée'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Vérification préalable : stock suffisant pour toutes les pièces ──
        for piece in facture.reparation.pieces.select_related('article_stock').all():
            if piece.article_stock and piece.article_stock.quantite < piece.quantite:
                return Response(
                    {'error': f"Stock insuffisant pour '{piece.article_stock.nom}' : "
                               f"{piece.article_stock.quantite} disponible(s), {piece.quantite} requise(s)."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        with transaction.atomic():
            # Génération sécurisée du numéro de facture (verrou en écriture)
            year = timezone.now().year
            last_invoice = (
                Facture.objects
                .filter(numero_facture__startswith=f"FAC-{year}")
                .select_for_update()
                .order_by('-numero_facture')
                .first()
            )

            if last_invoice and last_invoice.numero_facture:
                try:
                    last_num = int(last_invoice.numero_facture.split('-')[-1])
                    count = last_num + 1
                except (ValueError, IndexError):
                    count = Facture.objects.filter(type='DEFINITIVE').count() + 1
            else:
                count = Facture.objects.filter(type='DEFINITIVE').count() + 1

            facture.numero_facture = f"FAC-{year}-{count:04d}"
            facture.type = 'DEFINITIVE'
            facture.date_validation = timezone.now()
            facture.save()

            for piece in facture.reparation.pieces.select_related('article_stock').all():
                if piece.article_stock:
                    stock_item = piece.article_stock
                    qty_to_deduct = piece.quantite

                    stock_item.quantite = max(0, stock_item.quantite - qty_to_deduct)
                    stock_item.save()

                    MouvementStock.objects.create(
                        article=stock_item,
                        type_mouvement='SORTIE',
                        quantite=qty_to_deduct,
                        description=f"Sortie auto pour Facture {facture.numero_facture} (Réparation OR-{facture.reparation.id})",
                        utilisateur=request.user if request.user.is_authenticated else None
                    )

        return Response(FactureSerializer(facture).data)

    @action(detail=True, methods=['post'])
    def enregistrer_paiement(self, request, pk=None):
        facture = self.get_object()
        montant = Decimal(str(request.data.get('montant', 0)))
        
        # Cas spécial pour les factures à 0 (services gratuits ou erreurs)
        if facture.total_ttc <= 0:
            facture.statut_paiement = 'SOLDE'
            facture.save()
            return Response(FactureSerializer(facture).data)

        if montant <= 0 or montant > (facture.total_ttc - facture.montant_paye):
            return Response({'error': 'Montant invalide'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Règle des 75% pour accepter le paiement (Avance pour travaux)
        total_apres_paiement = facture.montant_paye + montant
        seuil_75_pourcent = facture.total_ttc * Decimal('0.75')
        
        if total_apres_paiement < seuil_75_pourcent:
            return Response({
                'error': f'Le paiement doit couvrir au minimum 75% du montant total ({seuil_75_pourcent:,.0f} FCFA) pour être accepté.'
            }, status=status.HTTP_400_BAD_REQUEST)

        facture.montant_paye += montant
        facture.mode_paiement = request.data.get('mode_paiement')
        facture.statut_paiement = 'SOLDE' if facture.montant_paye >= facture.total_ttc else 'PARTIEL'
        facture.save()
        
        # Si le paiement atteint au moins 75%, on peut entamer les réparations
        reparation = facture.reparation
        if reparation.statut == 'EN_ATTENTE':
            reparation.statut = 'EN_COURS'
            reparation.save()

        MouvementCaisse.objects.create(
            type_mouvement='RECETTE', 
            categorie='RECETTE_CLIENT', 
            montant=montant, 
            description=f"Paiement {facture.numero_facture}", 
            facture=facture,
            date_mouvement=timezone.now().date(),
            utilisateur=request.user if request.user.is_authenticated else None
        )

        # Notification Staff
        NotificationStaff.objects.create(
            type='PAIEMENT_RECU',
            message=f"Paiement de {montant:,.0f} F reçu pour la facture {facture.numero_facture}."
        )

        return Response(FactureSerializer(facture).data)

    @action(detail=True, methods=['post'])
    def relancer_paiement(self, request, pk=None):
        facture = self.get_object()
        client = facture.reparation.vehicule.client
        reste = facture.total_ttc - facture.montant_paye
        
        message = f"Rappel : Un solde de {reste:,.0f} F est en attente de paiement pour votre facture {facture.numero_facture or 'en cours'}. Merci de régulariser."
        
        NotificationClient.objects.create(
            client=client,
            type='FACTURE_ENVOYEE',
            message=message
        )
        
        return Response({'message': f'Relance envoyée à {client.nom}.'})

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        pdf = generate_document_pdf(self.get_object(), doc_type="FACTURE")
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Facture_{self.get_object().numero_facture}.pdf"'
        return response

    @action(detail=True, methods=['post'])
    def envoyer_email(self, request, pk=None):
        facture = self.get_object()
        client = facture.reparation.vehicule.client
        if not client.email:
            return Response({'error': 'Le client n\'a pas d\'adresse email.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pdf = generate_document_pdf(facture, doc_type="FACTURE")
            subject = f"Facture LUXEL-G - {facture.numero_facture or 'Proforma'}"
            body = f"Bonjour {client.nom} {client.prenoms},\n\nVeuillez trouver ci-joint votre facture concernant les travaux sur votre véhicule {facture.reparation.vehicule.immatriculation}.\n\nMontant Total : {facture.total_ttc} FCFA.\n\nCordialement,\nL'équipe LUXEL-G"
            
            email = EmailMessage(subject, body, to=[client.email])
            filename = f"Facture_{facture.numero_facture or 'Proforma'}.pdf"
            email.attach(filename, pdf, 'application/pdf')
            email.send()
            
            # Créer une notification client
            NotificationClient.objects.create(
                client=client,
                type='FACTURE_ENVOYEE', # Note: J'ai besoin d'ajouter ce type au modèle si nécessaire ou d'utiliser un existant
                message=f"Votre facture {facture.numero_facture or 'Proforma'} vous a été envoyée par email."
            )
            
            return Response({'message': 'Email envoyé avec succès !'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class DevisViewSet(viewsets.ModelViewSet):
    queryset = Devis.objects.all().order_by('-date_creation')
    serializer_class = DevisSerializer

    @action(detail=True, methods=['post'])
    def transformer_en_facture(self, request, pk=None):
        devis = self.get_object()
        facture, created = Facture.objects.get_or_create(reparation=devis.reparation, defaults={'type': 'PROFORMA', 'total_ht': devis.total_ht, 'total_ttc': devis.total_ttc})
        devis.statut = 'FACTURE'
        devis.save()
        return Response(FactureSerializer(facture).data)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        pdf = generate_document_pdf(self.get_object(), doc_type="DEVIS")
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Devis_{self.get_object().numero_devis}.pdf"'
        return response

    @action(detail=True, methods=['post'])
    def envoyer_email(self, request, pk=None):
        devis = self.get_object()
        client = devis.reparation.vehicule.client
        if not client.email:
            return Response({'error': 'Le client n\'a pas d\'adresse email.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pdf = generate_document_pdf(devis, doc_type="DEVIS")
            subject = f"Devis LUXEL-G - {devis.numero_devis or 'Brouillon'}"
            body = f"Bonjour {client.nom} {client.prenoms},\n\nVeuillez trouver ci-joint le devis concernant les travaux sur votre véhicule {devis.reparation.vehicule.immatriculation}.\n\nTotal Estimé : {devis.total_ttc} FCFA.\n\nCordialement,\nL'équipe LUXEL-G"
            
            email = EmailMessage(subject, body, to=[client.email])
            filename = f"Devis_{devis.numero_devis or 'Brouillon'}.pdf"
            email.attach(filename, pdf, 'application/pdf')
            email.send()
            
            # Créer une notification client
            NotificationClient.objects.create(
                client=client,
                type='DEVIS_ENVOYE',
                message=f"Le devis {devis.numero_devis or 'Brouillon'} concernant votre véhicule {devis.reparation.vehicule.immatriculation} vous a été envoyé par email."
            )
            
            return Response({'message': 'Email envoyé avec succès !'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ClientSpaceViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_permissions(self):
        if self.action in ['request_otp', 'verify_otp', 'register']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['post'])
    def request_otp(self, request):
        phone = request.data.get('phone')
        if not phone:
            return Response({'error': 'Numéro de téléphone requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        clean_phone = phone.replace(' ', '').replace('+', '')
        if clean_phone.startswith('229'): clean_phone = clean_phone[3:]
        
        client = Client.objects.filter(contact__icontains=clean_phone).first()
        if not client:
            return Response({'error': 'Client non trouvé. Veuillez contacter le garage.'}, status=status.HTTP_404_NOT_FOUND)
            
        if not client.email:
             return Response({'error': 'Aucune adresse email associée à votre compte. Veuillez contacter le garage pour mettre à jour vos informations.'}, status=status.HTTP_400_BAD_REQUEST)
        
        code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        print(f"\n=== [DEV] CODE DE CONNEXION OTP POUR {client.prenoms} {client.nom} ({phone}) : {code} ===\n")
        from .models import ClientOTP
        ClientOTP.objects.create(client=client, code=code)
        dev_otp_suffix = f" Code test : {code}" if settings.DEV_EXPOSE_OTP else ""
        
        try:
            subject = f"Votre code de connexion LUXEL-G : {code}"
            body = f"Bonjour {client.nom} {client.prenoms},\n\nVotre code de connexion pour l'espace client est : {code}\n\nCe code est valide pendant 10 minutes.\n\nCordialement,\nL'équipe LUXEL-G"
            email = EmailMessage(subject, body, to=[client.email])
            email.send()
            return Response({'message': f'Code envoyé par email.{dev_otp_suffix}'})
        except Exception as e:
            print(f"\n[WARNING] Échec de l'envoi de l'e-mail : {str(e)}\n")
            return Response({
                'message': f"Code généré ! L'envoi de l'e-mail a échoué.{dev_otp_suffix or ' Veuillez récupérer le code dans la console du serveur Django.'}"
            }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def verify_otp(self, request):
        phone = request.data.get('phone')
        code = request.data.get('code')
        
        if not phone or not code:
            return Response({'error': 'Numéro et code requis'}, status=status.HTTP_400_BAD_REQUEST)
            
        clean_phone = phone.replace(' ', '').replace('+', '')
        if clean_phone.startswith('229'): clean_phone = clean_phone[3:]
        
        client = Client.objects.filter(contact__icontains=clean_phone).first()
        if not client:
            return Response({'error': 'Client non trouvé'}, status=status.HTTP_404_NOT_FOUND)
            
        from .models import ClientOTP
        otp = ClientOTP.objects.filter(client=client, code=code, is_used=False).order_by('-created_at').first()
        
        if not otp or not otp.is_valid():
            return Response({'error': 'Code invalide ou expiré'}, status=status.HTTP_400_BAD_REQUEST)
            
        otp.is_used = True
        otp.save()
        
        if not client.user:
            username = f"client_{clean_phone}"
            user, created = User.objects.get_or_create(username=username, defaults={'email': client.email})
            client.user = user
            client.save()
        
        refresh = RefreshToken.for_user(client.user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'client_id': client.id
        })

    @action(detail=False, methods=['get'])
    def data(self, request):
        try:
            client = request.user.client_profile
        except AttributeError:
            return Response({'error': 'Profil client non trouvé pour cet utilisateur'}, status=status.HTTP_404_NOT_FOUND)
        
        vehicules = Vehicule.objects.filter(client=client)
        vehicules_data = MiniVehiculeSerializer(vehicules, many=True).data
        
        reparations = (
            Reparation.objects
            .filter(vehicule__client=client)
            .select_related('vehicule', 'vehicule__client', 'technicien')
            .prefetch_related('travaux', 'pieces', 'devis')
            .order_by('-date_creation')
        )
        reparations_data = ReparationSerializer(reparations, many=True).data
        
        factures = (
            Facture.objects
            .filter(reparation__vehicule__client=client)
            .select_related('reparation__vehicule__client')
            .order_by('-date_creation')
        )
        factures_data = FactureSerializer(factures, many=True).data
        
        # Calcul du solde impayé en base de données (plus performant)
        from django.db.models import Sum, F, ExpressionWrapper, DecimalField
        solde_result = factures.filter(type='DEFINITIVE').aggregate(
            solde=Sum(ExpressionWrapper(F('total_ttc') - F('montant_paye'), output_field=DecimalField()))
        )
        solde_impaye = solde_result['solde'] or Decimal('0')

        rdvs = Appointment.objects.filter(client=client).order_by('-date_rdv')
        rdvs_data = AppointmentSerializer(rdvs, many=True).data

        notifications = NotificationClient.objects.filter(client=client).order_by('-date_envoi')
        notifications_data = NotificationClientSerializer(notifications, many=True).data
        
        alertes = MaintenancePredictive.objects.filter(vehicule__client=client, actif=True)
        alertes_data = MaintenancePredictiveSerializer(alertes, many=True).data
        
        avis = Avis.objects.filter(client=client).order_by('-date_creation')
        avis_data = AvisSerializer(avis, many=True).data

        return Response({
            'client': ClientSerializer(client).data,
            'vehicules': vehicules_data,
            'reparations': reparations_data,
            'factures': factures_data,
            'rdvs': rdvs_data,
            'solde_impaye': solde_impaye,
            'notifications': notifications_data,
            'alertes': alertes_data,
            'avis': avis_data
        })

    @action(detail=False, methods=['post'])
    def submit_avis(self, request):
        try:
            client = request.user.client_profile
        except AttributeError:
            return Response({'error': 'Profil client non trouvé'}, status=status.HTTP_404_NOT_FOUND)
        
        data = {
            'client': client.id,
            'note': request.data.get('note'),
            'commentaire': request.data.get('commentaire'),
            'reparation': request.data.get('reparation')
        }
        
        serializer = AvisSerializer(data=data)
        if serializer.is_valid():
            avis = serializer.save()
            NotificationStaff.objects.create(
                type='NOUVEL_AVIS',
                message=f"Nouvel avis de {client.nom} ({avis.note}/5) : \"{avis.commentaire[:50]}...\""
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def register(self, request):
        data = request.data.copy()
        # Normalisation du numéro de téléphone
        phone = data.get('contact', '')
        clean_phone = phone.replace(' ', '').replace('+', '')
        if clean_phone.startswith('229'): clean_phone = clean_phone[3:]
        data['contact'] = clean_phone
        
        serializer = ClientSerializer(data=data)
        if serializer.is_valid():
            client = serializer.save()
            return Response(ClientSerializer(client).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='book-appointment')
    def book_appointment(self, request):
        try:
            client = request.user.client_profile
        except AttributeError:
            return Response({'error': 'Profil client non trouvé'}, status=status.HTTP_404_NOT_FOUND)
            
        data = request.data.copy()
        data['client'] = client.id
        
        serializer = AppointmentSerializer(data=data)
        if serializer.is_valid():
            appointment = serializer.save()
            # Notification Staff
            NotificationStaff.objects.create(
                type='NOUVEAU_RDV',
                message=f"Nouveau rendez-vous pris par {client.nom} {client.prenoms} pour le {appointment.date_rdv.strftime('%d/%m/%Y %H:%M')}."
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='download-invoice')
    def download_invoice_pdf(self, request):
        try:
            client = request.user.client_profile
        except AttributeError:
            return Response({'error': 'Accès non autorisé'}, status=status.HTTP_403_FORBIDDEN)

        invoice_id = request.query_params.get('invoice_id')
        if not invoice_id:
            return Response({'error': 'ID facture requis'}, status=status.HTTP_400_BAD_REQUEST)

        facture = Facture.objects.filter(id=invoice_id, reparation__vehicule__client=client).first()
        if not facture:
            return Response({'error': 'Facture non trouvée'}, status=status.HTTP_404_NOT_FOUND)

        pdf = generate_document_pdf(facture, doc_type="FACTURE")
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Facture_{facture.numero_facture or "Proforma"}.pdf"'
        return response

    @action(detail=False, methods=['post'], url_path='update')
    def update_profile(self, request):
        try:
            client = request.user.client_profile
        except AttributeError:
            return Response({'error': 'Profil client non trouvé'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ClientSerializer(client, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(ClientSerializer(client).data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class StockViewSet(viewsets.ModelViewSet):
    queryset = Stock.objects.all().order_by('nom')
    serializer_class = StockSerializer

    @action(detail=True, methods=['post'])
    def approvisionner(self, request, pk=None):
        item = self.get_object()
        qty = int(request.data.get('quantite', 0))
        item.quantite += qty
        item.save()
        
        # Créer le mouvement de stock
        MouvementStock.objects.create(
            article=item,
            type_mouvement='ENTREE',
            quantite=qty,
            description=request.data.get('description', 'Approvisionnement manuel'),
            utilisateur=request.user if request.user.is_authenticated else None
        )
        
        MouvementCaisse.objects.create(
            type_mouvement='DEPENSE', 
            categorie='ACHAT_PIECES', 
            montant=Decimal(str(request.data.get('prix_achat_total', 0))), 
            description=f"Achat {qty} x {item.nom}",
            date_mouvement=timezone.now().date(),
            utilisateur=request.user if request.user.is_authenticated else None
        )
        return Response(StockSerializer(item).data)

    @action(detail=True, methods=['post'])
    def ajuster_inventaire(self, request, pk=None):
        item = self.get_object()
        nouveau_physique = int(request.data.get('quantite_physique', item.quantite))
        theorique = item.stock_theorique
        ecart = nouveau_physique - theorique
        
        item.quantite = nouveau_physique
        item.save()
        
        MouvementStock.objects.create(
            article=item,
            type_mouvement='AJUSTEMENT',
            quantite=ecart,
            description=f"Ajustement Inventaire (Physique: {nouveau_physique}, Écart: {ecart})",
            utilisateur=request.user if request.user.is_authenticated else None
        )
        
        return Response(StockSerializer(item).data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Gestion Stock"
        
        # En-têtes identiques au fichier Luxury
        headers = ['Date', 'Réf', 'Désignation', 'Stock Initial', 'Entrées', 'Sorties', 'Stock Théorique', 'Stock Physique', 'Écart']
        ws.append(headers)
        
        # Style
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for item in self.get_queryset():
            ws.append([
                timezone.now().strftime('%d/%m/%Y'),
                item.sku,
                item.nom,
                item.stock_initial,
                item.entrees_total,
                item.sorties_total,
                item.stock_theorique,
                item.quantite,
                item.ecart
            ])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventaire_luxury_g.xlsx"'
        wb.save(response)
        return response

class MouvementCaisseViewSet(viewsets.ModelViewSet):
    queryset = MouvementCaisse.objects.all().order_by('-date_creation')
    serializer_class = MouvementCaisseSerializer

    @action(detail=False, methods=['get'])
    def synthese(self, request):
        qs = self.get_queryset()
        today = timezone.now().date()
        recettes = qs.filter(type_mouvement='RECETTE').aggregate(s=models.Sum('montant'))['s'] or 0
        depenses = qs.filter(type_mouvement='DEPENSE').aggregate(s=models.Sum('montant'))['s'] or 0
        r_j = qs.filter(type_mouvement='RECETTE', date_mouvement=today).aggregate(s=models.Sum('montant'))['s'] or 0
        d_j = qs.filter(type_mouvement='DEPENSE', date_mouvement=today).aggregate(s=models.Sum('montant'))['s'] or 0
        impayes = Facture.objects.filter(type='DEFINITIVE').aggregate(s=models.Sum(models.F('total_ttc') - models.F('montant_paye')))['s'] or 0
        return Response({'total_recettes': recettes, 'total_depenses': depenses, 'solde': recettes - depenses, 'recettes_jour': r_j, 'depenses_jour': d_j, 'total_impayes': impayes})

class MaintenancePredictiveViewSet(viewsets.ModelViewSet):
    queryset = MaintenancePredictive.objects.all().order_by('date_prochaine_prevue')
    serializer_class = MaintenancePredictiveSerializer

    @action(detail=False, methods=['get'])
    def alertes(self, request):
        alertes = self.queryset.filter(date_prochaine_prevue__lte=timezone.now() + timezone.timedelta(days=15), actif=True)
        return Response(MaintenancePredictiveSerializer(alertes, many=True).data)

class NotificationStaffViewSet(viewsets.ModelViewSet):
    queryset = NotificationStaff.objects.all().order_by('-date_creation')
    serializer_class = NotificationStaffSerializer

class LigneTravailViewSet(viewsets.ModelViewSet):
    queryset = LigneTravail.objects.all()
    serializer_class = LigneTravailSerializer

class LignePieceViewSet(viewsets.ModelViewSet):
    queryset = LignePiece.objects.all()
    serializer_class = LignePieceSerializer

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):
        if self.action == 'me':
            return [permissions.IsAuthenticated()]
        return [IsDirecteur()]

    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

class StatsViewSet(viewsets.ViewSet):
    permission_classes = [IsDirecteur]

    def list(self, request):
        # 1. Évolution du Revenu (6 derniers mois)
        six_months_ago = timezone.now() - timezone.timedelta(days=180)
        evolution_ca = MouvementCaisse.objects.filter(
            type_mouvement='RECETTE',
            date_mouvement__gte=six_months_ago
        ).annotate(
            month=TruncMonth('date_mouvement')
        ).values('month').annotate(
            total=models.Sum('montant')
        ).order_by('month')

        # 2. Top Catégories de Pannes
        top_pannes = Reparation.objects.values('categorie').annotate(
            count=models.Count('id')
        ).order_by('-count')[:5]

        # 3. Récapitulatif global Recettes/Dépenses
        flux = MouvementCaisse.objects.values('type_mouvement').annotate(
            total=models.Sum('montant')
        )

        # 4. KPIs (Compteurs)
        counts = {
            'clients': Client.objects.count(),
            'vehicles': Vehicule.objects.count(),
            'repairs_active': Reparation.objects.filter(statut='EN_COURS').count(),
            'stock_low': Stock.objects.filter(quantite__lt=models.F('seuil_alerte')).count()
        }

        return Response({
            'evolution_ca': list(evolution_ca),
            'top_pannes': list(top_pannes),
            'flux_global': list(flux),
            'counts': counts
        })
